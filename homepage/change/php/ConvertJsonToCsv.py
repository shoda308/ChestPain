# coding: UTF-8
# import sys
# import pprint

# pprint.pprint(sys.path)
from cmath import phase
import json
from re import I
from openpyxl import Workbook
import openpyxl as xl
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter
import openpyxl
# ワークブックの読み込み
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import sys
from pathlib import Path



args = sys.argv
filename = args[1]
dlfilename = args[2]

path = Path(__file__).parent   # Convert.JsonToCsv.pyのあるディレクトリ
path /= '../' + filename     # ディレクトリ移動

datalist = []
with open(path) as f:
    for line in f:
        
        datedata = line.split("::::")[0].split(',')[0]
        timedata = line.split("::::")[0].split(',')[1]
        numberdata = line.split("::::")[1]
        if '\n' in line:
            jsondatatext = line.split("::::")[2][:(len(line.split("::::")[2])-1)]
        jsondatatext = line.split("::::")[2]
        jsondata = json.loads(jsondatatext)
        
        templist = [datedata, timedata, numberdata, jsondata]
        if numberdata != 'XXXXXdeleted':
            datalist.append(templist)
        
# print(datalist)

# ここからエクセルを作成

# ワークブックの新規作成と保存
wb = Workbook()
wb.save(dlfilename)
wb.close()

wb = load_workbook(dlfilename)

# ワークシートの選択
ws = wb['Sheet']  # ワークシートを指定
ws.title = 'result'
ws = wb.active  # アクティブなワークシートを選択

# セルに書き込み
ws['A1'] = '日付'
ws['B1'] = '時刻'
ws['C1'] = '受付番号'
ws['D1'] = '痛みの種類'
ws['D2'] = '刺すような痛み'
ws['E2'] = '鈍い痛み'
ws['F2'] = '圧迫されるような痛み'
ws['G2'] = '締め付けるような痛み'
ws['H2'] = 'その他'
ws['I1'] = 'どこが痛い'
ws['I2'] = '左胸部'
ws['J2'] = '右胸部'
ws['K2'] = '前胸部'
ws['L2'] = '心窩部'
ws['M2'] = '背部'
ws['N2'] = '首や肩に放散'
ws['O2'] = '局所的'
ws['P2'] = 'その他'
ws['Q1'] = '持続時間'
ws['Q2'] = '瞬間的'
ws['R2'] = '5分以内'
ws['S2'] = '5分〜20分'
ws['T2'] = '20分以上'
ws['U2'] = 'それ以上'
ws['V1'] = 'どのような時'
ws['V2'] = '動いた時'
ws['W2'] = '安静時'
ws['X2'] = '体位を変えたとき'
ws['Y2'] = '呼吸との関係'
ws['Z2'] = '食中もしくは食後'
ws['AA2'] = '透析中'
ws['AB2'] = 'その他'
ws['AC1'] = '再現性はあるか'
ws['AC2'] = 'はい'
ws['AD2'] = 'いいえ'
ws['AE1'] = '時間帯'
ws['AE2'] = '朝方'
ws['AF2'] = '日中'
ws['AG2'] = '夜'
ws['AH2'] = '寝ているとき'
ws['AI2'] = '特にない'
ws['AJ2'] = 'その他'
ws['AK1'] = '他の症状'
ws['AK2'] = '動悸'
ws['AL2'] = '呼吸困難'
ws['AM2'] = '発熱'
ws['AN2'] = '冷汗'
ws['AO2'] = '吐き気、嘔吐'
ws['AP2'] = '倦怠感'
ws['AQ2'] = '失神・前失神'
ws['AR2'] = '特になし'
ws['AS2'] = 'その他'
ws['AT1'] = '性別'
ws['AT2'] = '男'
ws['AU2'] = '女'
ws['AV1'] = '世帯員数'
ws['AV2'] = '1人'
ws['AW2'] = '複数人'
ws['AX1'] = '信憑性'
ws['AX2'] = 'あり'
ws['AY2'] = 'なし'

wb.save(dlfilename)  # overwritedlfilename
wb.close()
# excelにデータを入力する

for idx in range(1, len(datalist)+1):#行指定
    #日付、時刻、受付番号入力
    for idxd in range(1, 4):
        ws.cell(row=idx+2,column=idxd).value = datalist[idx-1][idxd-1]
    #問診内容入力
    if  datalist[idx-1][3] != None:
        jsonObject = datalist[idx-1][3]
        if 'whatkind' in jsonObject:
            if 'A' in jsonObject["whatkind"]:
                ws.cell(row=idx+2,column=4).value = '1'
            if 'B' in jsonObject["whatkind"]:
                ws.cell(row=idx+2,column=5).value = '1'
            if 'C' in jsonObject["whatkind"]:
                ws.cell(row=idx+2,column=6).value = '1'
            if 'D' in jsonObject["whatkind"]:
                ws.cell(row=idx+2,column=7).value = '1'
            if 'H' in jsonObject["whatkind"]:
                if 'whatkindother' in jsonObject:
                    if jsonObject["whatkindother"] != '' and jsonObject["whatkindother"] != 'undefined':
                        ws.cell(row=idx+2,column=8).value = jsonObject["whatkindother"]
                    else:
                        ws.cell(row=idx+2,column=8).value = '1'
                else:
                    ws.cell(row=idx+2,column=8).value = '1'
        if 'where' in jsonObject:
            if 'A' in jsonObject["where"]:
                ws.cell(row=idx+2,column=9).value = '1'
            if 'B' in jsonObject["where"]:
                ws.cell(row=idx+2,column=10).value = '1'
            if 'C' in jsonObject["where"]:
                ws.cell(row=idx+2,column=11).value = '1'
            if 'D' in jsonObject["where"]:
                ws.cell(row=idx+2,column=12).value = '1'
            if 'E' in jsonObject["where"]:
                ws.cell(row=idx+2,column=13).value = '1'
            if 'F' in jsonObject["where"]:
                ws.cell(row=idx+2,column=14).value = '1'
            if 'G' in jsonObject["where"]:
                ws.cell(row=idx+2,column=15).value = '1'
            if 'H' in jsonObject["where"]:
                if 'otherplace' in jsonObject:
                    if jsonObject["otherplace"] != '' and jsonObject["otherplace"] != 'undefined':
                        ws.cell(row=idx+2,column=16).value = jsonObject["otherplace"]
                elif 'whereother' in jsonObject:
                    if jsonObject["whereother"] != '' and jsonObject["whereother"] != 'undefined':
                        ws.cell(row=idx+2,column=16).value = jsonObject["whereother"]
        if 'howlong' in jsonObject:
            if 'A' in jsonObject["howlong"]:
                ws.cell(row=idx+2,column=17).value = '1'
            if 'B' in jsonObject["howlong"]:
                ws.cell(row=idx+2,column=18).value = '1'
            if 'C' in jsonObject["howlong"]:
                ws.cell(row=idx+2,column=19).value = '1'
            if 'D' in jsonObject["howlong"]:
                ws.cell(row=idx+2,column=20).value = '1'
            if 'E' in jsonObject["howlong"]:
                ws.cell(row=idx+2,column=21).value = '1'
        if 'when' in jsonObject:
            if 'A' in jsonObject["when"]:
                ws.cell(row=idx+2,column=22).value = '1'
            if 'B' in jsonObject["when"]:
                ws.cell(row=idx+2,column=23).value = '1'
            if 'C' in jsonObject["when"]:
                ws.cell(row=idx+2,column=24).value = '1'
            if 'D' in jsonObject["when"]:
                ws.cell(row=idx+2,column=25).value = '1'
            if 'E' in jsonObject["when"]:
                ws.cell(row=idx+2,column=26).value = '1'
            if 'F' in jsonObject["when"]:
                ws.cell(row=idx+2,column=27).value = '1'
        if 'inaction' in jsonObject:
            if 'A' in jsonObject["inaction"]:
                ws.cell(row=idx+2,column=29).value = '1'
            if 'B' in jsonObject["inaction"]:
                ws.cell(row=idx+2,column=30).value = '1'
        if 'calm' in jsonObject:
            if 'A' in jsonObject["calm"]:
                ws.cell(row=idx+2,column=31).value = '1'
            if 'B' in jsonObject["calm"]:
                ws.cell(row=idx+2,column=32).value = '1'
            if 'C' in jsonObject["calm"]:
                ws.cell(row=idx+2,column=33).value = '1'
            if 'D' in jsonObject["calm"]:
                ws.cell(row=idx+2,column=34).value = '1'
            if 'E' in jsonObject["calm"]:
                ws.cell(row=idx+2,column=35).value = '1'
        if 'other' in jsonObject:
            if 'A' in jsonObject["other"]:
                ws.cell(row=idx+2,column=37).value = '1'
            if 'B' in jsonObject["other"]:
                ws.cell(row=idx+2,column=38).value = '1'
            if 'C' in jsonObject["other"]:
                ws.cell(row=idx+2,column=39).value = '1'
            if 'D' in jsonObject["other"]:
                ws.cell(row=idx+2,column=40).value = '1'
            if 'E' in jsonObject["other"]:
                ws.cell(row=idx+2,column=41).value = '1'
            if 'F' in jsonObject["other"]:
                ws.cell(row=idx+2,column=42).value = '1'
            if 'G' in jsonObject["other"]:
                ws.cell(row=idx+2,column=43).value = '1'
            if 'I' in jsonObject["other"]:
                ws.cell(row=idx+2,column=44).value = '1'
            if 'undefined' in jsonObject["other"]:
                ws.cell(row=idx+2,column=45).value = '1'
        if 'sex' in jsonObject:
            if 'A' in jsonObject["sex"]:
                ws.cell(row=idx+2,column=46).value = '1'
            if 'B' in jsonObject["sex"]:
                ws.cell(row=idx+2,column=47).value = '1'
        if 'constitution' in jsonObject:
            if 'A' in jsonObject["constitution"]:
                ws.cell(row=idx+2,column=48).value = '1'
            if 'B' in jsonObject["constitution"]:
                ws.cell(row=idx+2,column=49).value = '1'
        if 'clarity' in jsonObject:
            if 'A' in jsonObject["clarity"]:
                ws.cell(row=idx+2,column=50).value = '1'
            if 'B' in jsonObject["clarity"]:
                ws.cell(row=idx+2,column=51).value = '1'
        if 'other' not in jsonObject:
            ws.cell(row=idx+2,column=44).value = '1'
        if 'whenother' in jsonObject:
            if jsonObject["whenother"] != '' and jsonObject["whenother"] != 'undefined':
                ws.cell(row=idx+2,column=28).value = jsonObject["whenother"]
        if 'calmother' in jsonObject:
            if jsonObject["calmother"] != '' and jsonObject["calmother"] != 'undefined':
                ws.cell(row=idx+2,column=36).value = jsonObject["calmother"]
        if 'otherother' in jsonObject: 
            if jsonObject["otherother"] != '' and jsonObject["otherother"] != 'undefined':
                ws.cell(row=idx+2,column=45).value = jsonObject["otherother"]
             
    
    
        
    wb.save(dlfilename)
    wb.close()
    

#ここから列幅調整

# set input file name
inputfile = dlfilename

# read input xlsx
wb1 = xl.load_workbook(filename=inputfile)
ws1 = wb1.worksheets[0]

# set column width
for col in ws1.columns:
    max_length = 0
    column = col[0].column

    for cell in col:   
        if len(str(cell.value)) > max_length:
            max_length = len(str(cell.value))
    
    adjusted_width = (max_length + 1) * 1.8
    ws1.column_dimensions[get_column_letter(column)].width = adjusted_width

wb1['result'].column_dimensions['A'].width = 14
wb1['result'].column_dimensions['B'].width = 11
wb1['result'].column_dimensions['C'].width = 13

# save xlsx file
wb1.save(inputfile)
wb.close()






#ここから羅線

side = Side(style='thin', color='000000')

# set border (black thin line)
border = Border(top=side, bottom=side, left=side, right=side)

# write in sheet
for row in ws1:
    for cell in row:
        ws1[cell.coordinate].border = border

# save xlsx file
wb1.save(inputfile)


#上下中央揃え
ws1['A1'].alignment =Alignment(horizontal = 'center',  vertical = 'center',wrap_text = False)
ws1['B1'].alignment =Alignment(horizontal = 'center',  vertical = 'center',wrap_text = False)
ws1['C1'].alignment =Alignment(horizontal = 'center',  vertical = 'center',wrap_text = False)

#右寄せ
mylist3=[]
i = 2
for w in datalist:
    i = i + 1
    mylist3.append(i)
    

for list in mylist3:
    for row in ws.iter_rows():
        for cell in row:
            if cell.row == list:
                cell.alignment =Alignment(horizontal = 'right')

wb1.save(inputfile)
wb.close()


#ここからしましま
wb = openpyxl.load_workbook(dlfilename)
ws = wb['result']


mylist=[]
mylist2=[1,2]
i = 2
for w in datalist:
    i = i + 1
    if i % 2 == 0:
        mylist.append(i)
    

for list in mylist:
    for row in ws.iter_rows():
        for cell in row:
            if cell.row == list:
                cell.fill = PatternFill(fgColor='E6E6FA',bgColor="E6E6FA", fill_type = "solid")

for list in mylist2:
    for row in ws.iter_rows():
        for cell in row:
            if cell.row == list:
                cell.fill = PatternFill(patternType='solid', start_color='B0C4DE', end_color='B0C4DE')
#別名で保存
wb.save(dlfilename)
wb.close()

#ここからセルの結合

# ブックを取得
book = openpyxl.load_workbook(dlfilename)
# シートを取得
sheet = book['result']

# セルを結合
sheet.merge_cells('D1:H1')
sheet.merge_cells('I1:P1')
sheet.merge_cells('Q1:U1')
sheet.merge_cells('V1:AB1')
sheet.merge_cells('AC1:AD1')
sheet.merge_cells('AE1:AJ1')
sheet.merge_cells('AK1:AS1')
sheet.merge_cells('AT1:AU1')
sheet.merge_cells('AV1:AW1')
sheet.merge_cells('AX1:AY1')
sheet.merge_cells('A1:A2')
sheet.merge_cells('B1:B2')
sheet.merge_cells('C1:C2')

# 保存する
book.save(dlfilename)
wb.close()


#ここから行固定

wb = load_workbook(dlfilename)
ws = wb.active

ws.freeze_panes = 'D3'

wb.save(dlfilename)
wb.close()
