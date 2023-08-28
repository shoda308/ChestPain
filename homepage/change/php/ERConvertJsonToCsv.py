# coding: UTF-8
# import sys
# import pprint

# pprint.pprint(sys.path)
from cmath import phase
import json
from re import I
from openpyxl import Workbook
import openpyxl as xl
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter
import openpyxl
# ワークブックの読み込み
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import sys
from pathlib import Path



args = sys.argv
filename = args[1]
dlfilename = args[2]

path = Path(__file__).parent   # Convert.JsonToCsv.pyのあるディレクトリ
path /= '../' + filename     # ディレクトリ移動

datalist = []
with open(path) as f:
    for line in f:
        
        datedata = line.split("::::")[0].split(',')[0]
        timedata = line.split("::::")[0].split(',')[1]
        numberdata = line.split("::::")[1]
        if '\n' in line:
            jsondatatext = line.split("::::")[2][:(len(line.split("::::")[2])-1)]
        jsondatatext = line.split("::::")[2]
        jsondata = json.loads(jsondatatext)
        
        templist = [datedata, timedata, numberdata, jsondata]
        if numberdata != 'XXXXXdeleted':
            datalist.append(templist)
        
# print(datalist)

# ここからエクセルを作成

# ワークブックの新規作成と保存
wb = Workbook()
wb.save(dlfilename)
wb.close()

wb = load_workbook(dlfilename)

# ワークシートの選択
ws = wb['Sheet']  # ワークシートを指定
ws.title = 'result'
ws = wb.active  # アクティブなワークシートを選択

# セルに書き込み
calms2 = ['刺すような痛み','鈍い痛み','圧迫されるような痛み','締め付けるような痛み','その他','左胸部','右胸部','前胸部',
         '心窩部','背部','首や肩や歯に放散','局所的（指で刺せるような）','その他','動いた時(労作)','安静時','体位を変えた時',
         '呼吸時(深呼吸など)','食中もしくは食後','その他','朝方','日中','夜','寝ているとき','特にない','その他','動悸',
         '呼吸困難','発熱','冷汗','吐き気、おう吐','倦怠感','意識喪失','特にない','その他','はい','いいえ','48時間以内',
         '1ヶ月以内','2ヶ月以内','A症状の闘値が低下','B安静時にも生じるようになった','C症状の頻度増加','D症状の性質変化',
         'E症状が強くなった','F症状が弱くなった','G症状の部位変化','H症状の持続時間増加','I症状の持続時間低下','J随伴症状変化',
         '左胸部','右胸部','前胸部','心窩部','背部','首や肩や歯に放散','局所的（指で刺せるような）','その他','瞬間的(数秒以内)',
         '5分以内','5~20分以内','20分以上~数時間','それ以上','動悸','呼吸困難','発熱','冷汗','吐き気、おう吐','倦怠感','意識喪失','その他']
for i, w in enumerate(calms2):
    ws.cell(row=2,column=i+4).value = w
ws['A1'] = '日付'
ws['B1'] = '時刻'
ws['C1'] = '受付番号'
ws['D1'] = '胸痛の症状'
ws['I1'] = '症状の部位'
ws['Q1'] = '何をしているとき'
ws['W1'] = '時間帯'
ws['AC1'] = '随伴症状'
ws['AL1'] = '初めての症状か'
ws['AN1'] = '前回の起きたのはいつ'
ws['AQ1'] = 'どのように変化した'
ws['BA1'] = '場所変化'
ws['BI1'] = '持続時間変化'
ws['BN1'] = '随伴症状変化'

#ここから列幅調整
# set column width
for col in ws.columns:
    max_length = 0
    column = col[0].column

    for cell in col:   
        if len(str(cell.value)) > max_length:
            max_length = len(str(cell.value))
    
    adjusted_width = (max_length + 1) * 1.8
    ws.column_dimensions[get_column_letter(column)].width = adjusted_width

wb['result'].column_dimensions['A'].width = 14
wb['result'].column_dimensions['B'].width = 11
wb['result'].column_dimensions['C'].width = 13

#ここから羅線

side = Side(style='thin', color='000000')

# set border (black thin line)
border = Border(top=side, bottom=side, left=side, right=side)

# write in sheet
for row in ws:
    for cell in row:
        ws[cell.coordinate].border = border

# save xlsx file
wb.save(dlfilename)


#上下中央揃え
ws['A1'].alignment =Alignment(horizontal = 'center',  vertical = 'center',wrap_text = False)
ws['B1'].alignment =Alignment(horizontal = 'center',  vertical = 'center',wrap_text = False)
ws['C1'].alignment =Alignment(horizontal = 'center',  vertical = 'center',wrap_text = False)

#右寄せ
mylist3=[]
i = 2
for w in datalist:
    i = i + 1
    mylist3.append(i)
    

for list in mylist3:
    for row in ws.iter_rows():
        for cell in row:
            if cell.row == list:
                cell.alignment =Alignment(horizontal = 'right')


wb.save(dlfilename)  # overwritedlfilename
wb.close()
# excelにデータを入力する

# for idx in range(1, len(datalist)+1):#行指定
#     #日付、時刻、受付番号入力
#     for idxd in range(1, 4):
#         ws.cell(row=idx+2,column=idxd).value = datalist[idx-1][idxd-1]
#     #問診内容入力
#     if  datalist[idx-1][3] != None:
#         jsonObject = datalist[idx-1][3]
#         if 'whatkind' in jsonObject:
#             if 'A' in jsonObject["whatkind"]:
#                 ws.cell(row=idx+2,column=4).value = '○'
#             if 'B' in jsonObject["whatkind"]:
#                 ws.cell(row=idx+2,column=5).value = '○'
#             if 'C' in jsonObject["whatkind"]:
#                 ws.cell(row=idx+2,column=6).value = '○'
#             if 'D' in jsonObject["whatkind"]:
#                 ws.cell(row=idx+2,column=7).value = '○'
#             if 'H' in jsonObject["whatkind"]:
#                 if 'whatkindother' in jsonObject:
#                     if jsonObject["whatkindother"] != '' and jsonObject["whatkindother"] != 'undefined':
#                         ws.cell(row=idx+2,column=8).value = jsonObject["whatkindother"]
#                     else:
#                         ws.cell(row=idx+2,column=8).value = '○'
#                 else:
#                     ws.cell(row=idx+2,column=8).value = '○'
#         if 'where' in jsonObject:
#             if 'A' in jsonObject["where"]:
#                 ws.cell(row=idx+2,column=9).value = '○'
#             if 'B' in jsonObject["where"]:
#                 ws.cell(row=idx+2,column=10).value = '○'
#             if 'C' in jsonObject["where"]:
#                 ws.cell(row=idx+2,column=11).value = '○'
#             if 'E' in jsonObject["where"]:
#                 ws.cell(row=idx+2,column=12).value = '○'
#             if 'G' in jsonObject["where"]:
#                 ws.cell(row=idx+2,column=13).value = '○'
#             if 'H' in jsonObject["where"]:
#                 if 'otherplace' in jsonObject:
#                     if jsonObject["otherplace"] != '' and jsonObject["otherplace"] != 'undefined':
#                         ws.cell(row=idx+2,column=14).value = jsonObject["otherplace"]
#                 elif 'whereother' in jsonObject:
#                     if jsonObject["whereother"] != '' and jsonObject["whereother"] != 'undefined':
#                         ws.cell(row=idx+2,column=14).value = jsonObject["whereother"]
#         if 'howlong' in jsonObject:
#             if 'A' in jsonObject["howlong"]:
#                 ws.cell(row=idx+2,column=15).value = '○'
#             if 'B' in jsonObject["howlong"]:
#                 ws.cell(row=idx+2,column=16).value = '○'
#             if 'C' in jsonObject["howlong"]:
#                 ws.cell(row=idx+2,column=17).value = '○'
#             if 'D' in jsonObject["howlong"]:
#                 ws.cell(row=idx+2,column=18).value = '○'
#             if 'E' in jsonObject["howlong"]:
#                 ws.cell(row=idx+2,column=19).value = '○'
#         if 'when' in jsonObject:
#             if 'A' in jsonObject["when"]:
#                 ws.cell(row=idx+2,column=20).value = '○'
#             if 'B' in jsonObject["when"]:
#                 ws.cell(row=idx+2,column=21).value = '○'
#             if 'C' in jsonObject["when"]:
#                 ws.cell(row=idx+2,column=22).value = '○'
#             if 'D' in jsonObject["when"]:
#                 ws.cell(row=idx+2,column=23).value = '○'
#             if 'E' in jsonObject["when"]:
#                 ws.cell(row=idx+2,column=24).value = '○'
#             if 'F' in jsonObject["when"]:
#                 ws.cell(row=idx+2,column=25).value = '○'
#         if 'inaction' in jsonObject:
#             if 'A' in jsonObject["inaction"]:
#                 ws.cell(row=idx+2,column=27).value = '○'
#             if 'B' in jsonObject["inaction"]:
#                 ws.cell(row=idx+2,column=28).value = '○'
#         if 'calm' in jsonObject:
#             if 'A' in jsonObject["calm"]:
#                 ws.cell(row=idx+2,column=29).value = '○'
#             if 'B' in jsonObject["calm"]:
#                 ws.cell(row=idx+2,column=30).value = '○'
#             if 'C' in jsonObject["calm"]:
#                 ws.cell(row=idx+2,column=31).value = '○'
#             if 'D' in jsonObject["calm"]:
#                 ws.cell(row=idx+2,column=32).value = '○'
#             if 'E' in jsonObject["calm"]:
#                 ws.cell(row=idx+2,column=33).value = '○'
#         if 'other' in jsonObject:
#             if 'A' in jsonObject["other"]:
#                 ws.cell(row=idx+2,column=35).value = '○'
#             if 'B' in jsonObject["other"]:
#                 ws.cell(row=idx+2,column=36).value = '○'
#             if 'C' in jsonObject["other"]:
#                 ws.cell(row=idx+2,column=37).value = '○'
#             if 'D' in jsonObject["other"]:
#                 ws.cell(row=idx+2,column=38).value = '○'
#             if 'E' in jsonObject["other"]:
#                 ws.cell(row=idx+2,column=39).value = '○'
#             if 'F' in jsonObject["other"]:
#                 ws.cell(row=idx+2,column=40).value = '○'
#             if 'G' in jsonObject["other"]:
#                 ws.cell(row=idx+2,column=41).value = '○'
#         if 'whenother' in jsonObject:
#             if jsonObject["whenother"] != '' and jsonObject["whenother"] != 'undefined':
#                 ws.cell(row=idx+2,column=26).value = jsonObject["whenother"]
#         if 'calmother' in jsonObject:
#             if jsonObject["calmother"] != '' and jsonObject["calmother"] != 'undefined':
#                 ws.cell(row=idx+2,column=34).value = jsonObject["calmother"]
#         if 'otherother' in jsonObject: 
#             if jsonObject["otherother"] != '' and jsonObject["otherother"] != 'undefined':
#                 ws.cell(row=idx+2,column=42).value = jsonObject["otherother"]
             
    
    
        
#     wb.save(dlfilename)
#     wb.close()
    











# #ここからしましま

# mylist=[]
# mylist2=[1,2]
# i = 2
# for w in datalist:
#     i = i + 1
#     if i % 2 == 0:
#         mylist.append(i)
    

# for list in mylist:
#     for row in ws.iter_rows():
#         for cell in row:
#             if cell.row == list:
#                 cell.fill = PatternFill(fgColor='E6E6FA',bgColor="E6E6FA", fill_type = "solid")

# for list in mylist2:
#     for row in ws.iter_rows():
#         for cell in row:
#             if cell.row == list:
#                 cell.fill = PatternFill(patternType='solid', start_color='B0C4DE', end_color='B0C4DE')
# #別名で保存
# wb.save(dlfilename)
# wb.close()

# #ここからセルの結合

# # ブックを取得
# book = openpyxl.load_workbook(dlfilename)
# # シートを取得
# sheet = book['result']

# # セルを結合
# sheet.merge_cells('D1:H1')
# sheet.merge_cells('I1:N1')
# sheet.merge_cells('O1:S1')
# sheet.merge_cells('T1:Z1')
# sheet.merge_cells('AA1:AB1')
# sheet.merge_cells('AC1:AH1')
# sheet.merge_cells('AI1:AP1')
# sheet.merge_cells('A1:A2')
# sheet.merge_cells('B1:B2')
# sheet.merge_cells('C1:C2')

# # 保存する
# book.save(dlfilename)
# wb.close()


# #ここから行固定

# wb = load_workbook(dlfilename)
# ws = wb.active

# ws.freeze_panes = 'D3'

# wb.save(dlfilename)
# wb.close()